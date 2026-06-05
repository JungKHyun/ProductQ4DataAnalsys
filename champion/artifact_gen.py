"""산출물 자동 생성 모듈 (HTML · Excel · PNG)"""
from __future__ import annotations
import io
import json
import re

import matplotlib
matplotlib.use("Agg")
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from google import genai
from google.genai import types

from .profiler import DataProfile
from .problem_gen import Problem


# ── 한글 폰트 설정 ────────────────────────────────────────────
def _set_font():
    for name in ["Malgun Gothic", "NanumGothic", "AppleGothic", "DejaVu Sans"]:
        if any(f.name == name for f in fm.fontManager.ttflist):
            plt.rcParams["font.family"] = name
            break
    plt.rcParams["axes.unicode_minus"] = False

_set_font()
plt.show = lambda *a, **k: None


# ── HTML 생성 공통 LLM 호출 ──────────────────────────────────
_HTML_SYSTEM = """당신은 데이터 시각화 전문 웹 개발자입니다.
주어진 데이터와 요구사항을 바탕으로 완전히 동작하는 단일 HTML 파일을 생성합니다.

규칙:
- 순수 HTML + Vanilla JS만 사용 (React/Vue 금지)
- 외부 라이브러리는 CDN으로 포함 (Chart.js: https://cdn.jsdelivr.net/npm/chart.js)
- 데이터는 HTML 내부 <script> const DATA = [...]; 형태로 임베드
- 응답은 <!DOCTYPE html>로 시작하는 완전한 HTML 코드만 출력 (코드블록 없이)"""


def _llm_html(client: genai.Client, model: str, prompt: str) -> str:
    resp = client.models.generate_content(
        model=model,
        contents=prompt,
        config=types.GenerateContentConfig(
            system_instruction=_HTML_SYSTEM,
            temperature=0.3,
            max_output_tokens=8192,
        ),
    )
    raw = (resp.text or "").strip()
    raw = re.sub(r"^```html?\s*", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s*```$", "", raw)
    idx = raw.find("<!DOCTYPE")
    if idx == -1:
        idx = raw.find("<html")
    if idx > 0:
        raw = raw[idx:]
    return raw


def _df_to_json(df: pd.DataFrame, max_rows: int = 500) -> str:
    sample = df.head(max_rows).copy()
    for col in sample.select_dtypes(include=["datetime64"]).columns:
        sample[col] = sample[col].astype(str)
    return json.dumps(sample.to_dict(orient="records"), ensure_ascii=False, default=str)


# ── 1. HTML 대시보드 ─────────────────────────────────────────
def generate_html_dashboard(
    client: genai.Client,
    model: str,
    df: pd.DataFrame,
    profile: DataProfile,
    problem: Problem,
) -> bytes:
    data_json = _df_to_json(df)
    num_cols = profile.numeric_cols[:3]
    cat_cols = profile.categorical_cols[:2]

    prompt = f"""다음 데이터를 활용해 아래 요구사항을 만족하는 HTML 대시보드를 제작하라.

[시나리오]
{problem.scenario}

[필수 기능]
{chr(10).join(f"- {r}" for r in problem.requirements)}

[데이터 구조]
컬럼: {", ".join(df.columns.tolist())}
수치형: {", ".join(num_cols)}
범주형: {", ".join(cat_cols)}
총 {len(df)}행

[데이터 JSON (최대 500행, <script>에 임베드)]
{data_json[:6000]}

[기술 요건]
- Chart.js CDN으로 차트 2개 이상 (막대 + 파이 또는 선)
- KPI 카드 3개 이상 (총계, 평균 등 실제 계산값)
- 범주형 필터 드롭다운 (onChange 이벤트로 차트 업데이트)
- 자동 요약 문장 생성 (JS로 계산하여 화면 출력)
- 반응형 레이아웃 (CSS Grid)
- 공공기관 스타일 (파란색 #1a3a5c 포인트, 흰 배경)"""

    return _llm_html(client, model, prompt).encode("utf-8")


# ── 2. HTML 검색/필터 서비스 ─────────────────────────────────
def generate_html_filter(
    client: genai.Client,
    model: str,
    df: pd.DataFrame,
    profile: DataProfile,
    problem: Problem,
) -> bytes:
    data_json = _df_to_json(df)

    prompt = f"""다음 데이터를 활용해 검색·필터 서비스 HTML 파일을 제작하라.

[시나리오]
{problem.scenario}

[필수 기능]
{chr(10).join(f"- {r}" for r in problem.requirements)}

[데이터 구조]
컬럼: {", ".join(df.columns.tolist())}
범주형 (필터용): {", ".join(profile.categorical_cols[:4])}
총 {len(df)}행

[데이터 JSON]
{data_json[:6000]}

[기술 요건]
- 상단 검색창: 실시간 텍스트 필터 (keyup 이벤트)
- 드롭다운 필터: 범주형 컬럼별 각 1개 (동적 옵션 생성)
- 결과 테이블: 헤더 클릭 정렬, 줄무늬 스타일
- 결과 건수 실시간 표시 ("총 OOO건 중 OO건")
- 테이블 페이지네이션 (한 페이지 20행)
- 전체 초기화 버튼
- 공공기관 스타일"""

    return _llm_html(client, model, prompt).encode("utf-8")


# ── 3. Excel + VBA ────────────────────────────────────────────
def generate_excel_vba(
    client: genai.Client,
    model: str,
    df: pd.DataFrame,
    profile: DataProfile,
    problem: Problem,
) -> tuple[bytes, str]:
    """
    Returns:
        xlsx_bytes: 데이터가 담긴 Excel 파일 (xlsx 형식)
        vba_code:   VBA 매크로 코드 문자열 (.bas 파일용)
    """
    wb = openpyxl.Workbook()

    # ── 원본데이터 시트 ────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "원본데이터"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for ci, col_name in enumerate(df.columns, 1):
        cell = ws_data.cell(1, ci, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws_data.column_dimensions[get_column_letter(ci)].width = max(len(str(col_name)) + 4, 12)

    for ri, row in enumerate(df.head(1000).itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws_data.cell(ri, ci, val if pd.notna(val) else "")

    # ── 집계요약 시트 ──────────────────────────────────────
    ws_sum = wb.create_sheet("집계요약")
    ws_sum["A1"] = "집계 항목"
    ws_sum["B1"] = "값"
    for c in ["A1", "B1"]:
        ws_sum[c].fill = header_fill
        ws_sum[c].font = header_font
        ws_sum[c].alignment = Alignment(horizontal="center")

    agg_rows = [("전체 데이터 건수", len(df))]
    for num_col in profile.numeric_cols[:4]:
        agg_rows += [
            (f"{num_col} 합계", round(float(df[num_col].sum()), 2)),
            (f"{num_col} 평균", round(float(df[num_col].mean()), 2)),
            (f"{num_col} 최대", round(float(df[num_col].max()), 2)),
        ]
    for cat_col in profile.categorical_cols[:2]:
        mode_val = df[cat_col].mode()
        agg_rows.append((f"{cat_col} 최빈값", str(mode_val.iloc[0]) if len(mode_val) else ""))

    for ri, (label, val) in enumerate(agg_rows, 2):
        ws_sum.cell(ri, 1, label).alignment = Alignment(horizontal="left")
        ws_sum.cell(ri, 2, val).alignment = Alignment(horizontal="right")

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 18

    # 차트 추가
    if len(agg_rows) > 1:
        chart = BarChart()
        chart.type = "col"
        chart.title = "주요 집계 현황"
        chart.y_axis.title = "값"
        chart.grouping = "clustered"
        num_data_count = sum(1 for _, v in agg_rows if isinstance(v, (int, float)))
        if num_data_count > 0:
            data_ref = Reference(ws_sum, min_col=2, min_row=2, max_row=1 + len(agg_rows))
            cats_ref = Reference(ws_sum, min_col=1, min_row=2, max_row=1 + len(agg_rows))
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cats_ref)
            chart.width = 20
            chart.height = 12
            ws_sum.add_chart(chart, "D2")

    # ── 자동보고서 시트 (VBA가 채울 영역) ─────────────────
    ws_rep = wb.create_sheet("자동보고서")
    ws_rep["A1"] = "📋 이 시트는 VBA 매크로 실행 후 자동으로 채워집니다."
    ws_rep["A2"] = "→ Excel 개발 도구 탭 > 매크로 > '자동보고서생성' 실행"
    ws_rep["A1"].font = Font(size=12, bold=True, color="1F4E79")
    ws_rep["A2"].font = Font(size=11, italic=True, color="666666")

    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)
    xlsx_bytes = xlsx_buf.read()

    # ── VBA 코드 생성 (LLM) ───────────────────────────────
    vba_prompt = f"""다음 Excel 파일 구조에 맞는 VBA 매크로를 작성하라.

[Excel 시트 구조]
- "원본데이터" 시트: {", ".join(df.columns.tolist())} 컬럼 ({len(df)}행)
- "집계요약" 시트: 집계된 통계값
- "자동보고서" 시트: VBA로 내용을 채워야 함

[시나리오]
{problem.scenario}

[필수 구현 사항]
{chr(10).join(f"- {r}" for r in problem.requirements)}

[VBA 요구사항]
1. Sub 이름: 자동보고서생성()
2. "원본데이터" 시트에서 데이터 읽기
3. "자동보고서" 시트에 다음 내용 작성:
   - 보고서 제목 (대형 폰트, 색상 적용)
   - 작성 일시 (Now() 함수)
   - 수치형 컬럼({", ".join(profile.numeric_cols[:3])}) 기준 집계표
   - 범주형 컬럼({", ".join(profile.categorical_cols[:2])}) 기준 그룹별 집계
   - 요약 텍스트
4. 셀 서식: 헤더 배경색, 테두리, 볼드체
5. 완료 시 MsgBox "보고서 생성 완료!" 표시

Sub 자동보고서생성()부터 End Sub까지 VBA 코드만 출력:"""

    vba_resp = client.models.generate_content(
        model=model,
        contents=vba_prompt,
        config=types.GenerateContentConfig(temperature=0.3, max_output_tokens=3000),
    )
    vba_raw = (vba_resp.text or "").strip()
    vba_raw = re.sub(r"^```vba?\s*", "", vba_raw, flags=re.IGNORECASE)
    vba_raw = re.sub(r"\s*```$", "", vba_raw)

    return xlsx_bytes, vba_raw


# ── 4. PNG 인포그래픽 ─────────────────────────────────────────
def generate_png_infographic(
    df: pd.DataFrame,
    profile: DataProfile,
    problem: Problem,
) -> bytes:
    BG = "#0d1117"
    ACCENT = "#58a6ff"
    SECONDARY = "#f78166"
    TEXT = "#e6edf3"
    CARD_BG = "#161b22"
    BORDER = "#30363d"

    fig = plt.figure(figsize=(18, 11), facecolor=BG)

    # 그리드: 2행 3열
    gs = fig.add_gridspec(
        3, 3, hspace=0.45, wspace=0.35,
        left=0.04, right=0.96, top=0.90, bottom=0.06,
    )

    # ── 타이틀 ────────────────────────────────────────────
    fig.text(0.5, 0.95, problem.title, ha="center", va="top",
             fontsize=22, fontweight="bold", color=TEXT)
    fig.text(0.5, 0.915, problem.scenario[:90] + ("…" if len(problem.scenario) > 90 else ""),
             ha="center", va="top", fontsize=9, color="#8b949e")

    # ── KPI 카드 (상단 3개) ───────────────────────────────
    kpi_items = []
    for num_col in profile.numeric_cols[:3]:
        total = df[num_col].sum()
        kpi_items.append((f"총 {num_col}", f"{total:,.0f}"))
    if len(kpi_items) < 3 and profile.numeric_cols:
        col = profile.numeric_cols[0]
        extras = [
            (f"평균 {col}", f"{df[col].mean():.2f}"),
            (f"최대 {col}", f"{df[col].max():,.0f}"),
            ("전체 건수", f"{len(df):,}"),
        ]
        for e in extras:
            if len(kpi_items) >= 3:
                break
            if e not in kpi_items:
                kpi_items.append(e)
    kpi_items = kpi_items[:3]

    for i, (label, val) in enumerate(kpi_items):
        ax = fig.add_subplot(gs[0, i])
        ax.set_facecolor(CARD_BG)
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis("off")
        rect = plt.Rectangle((0.02, 0.02), 0.96, 0.96,
                               fill=True, facecolor=CARD_BG,
                               edgecolor=ACCENT if i == 0 else BORDER, linewidth=2,
                               transform=ax.transAxes)
        ax.add_patch(rect)
        ax.text(0.5, 0.62, val, ha="center", va="center",
                fontsize=20, fontweight="bold",
                color=ACCENT if i == 0 else SECONDARY,
                transform=ax.transAxes)
        ax.text(0.5, 0.28, label, ha="center", va="center",
                fontsize=10, color=TEXT, transform=ax.transAxes)

    # ── 차트 1: 수평 막대 (좌측 하단) ───────────────────────
    ax1 = fig.add_subplot(gs[1:, :2])
    ax1.set_facecolor(CARD_BG)
    if profile.categorical_cols and profile.numeric_cols:
        cat_col = profile.categorical_cols[0]
        num_col = profile.numeric_cols[0]
        grp = df.groupby(cat_col)[num_col].sum().nlargest(10)
        colors_bar = [ACCENT if i == 0 else "#1f6feb" for i in range(len(grp))]
        bars = ax1.barh(range(len(grp)), grp.values, color=colors_bar, height=0.65)
        ax1.set_yticks(range(len(grp)))
        ax1.set_yticklabels(
            [str(v)[:15] for v in grp.index.tolist()],
            color=TEXT, fontsize=9,
        )
        ax1.set_xlabel(num_col, color="#8b949e", fontsize=9)
        ax1.set_title(
            f"{cat_col}별 {num_col} 상위 10",
            color=TEXT, fontsize=12, pad=10, fontweight="bold",
        )
        ax1.tick_params(colors="#8b949e", labelsize=8)
        for sp in ax1.spines.values():
            sp.set_color(BORDER)
        ax1.xaxis.label.set_color("#8b949e")
        for bar, v in zip(bars, grp.values):
            ax1.text(v * 0.99, bar.get_y() + bar.get_height() / 2,
                     f"{v:,.0f}", ha="right", va="center",
                     color=TEXT, fontsize=8)
    else:
        ax1.text(0.5, 0.5, "데이터 없음", ha="center", va="center", color=TEXT)
    ax1.set_facecolor(CARD_BG)

    # ── 차트 2: 도넛 (우측 하단) ──────────────────────────
    ax2 = fig.add_subplot(gs[1:, 2])
    ax2.set_facecolor(CARD_BG)
    if profile.categorical_cols and profile.numeric_cols:
        cat_col = profile.categorical_cols[0]
        num_col = profile.numeric_cols[0]
        grp = df.groupby(cat_col)[num_col].sum().nlargest(6)
        donut_colors = [ACCENT, SECONDARY, "#3fb950", "#d29922", "#a5d6ff", "#f2a097"]
        wedges, _, autos = ax2.pie(
            grp.values,
            autopct="%1.1f%%",
            pctdistance=0.78,
            colors=donut_colors[:len(grp)],
            startangle=90,
            wedgeprops={"width": 0.52, "edgecolor": BG, "linewidth": 2},
        )
        for at in autos:
            at.set_fontsize(8)
            at.set_color(TEXT)
        ax2.set_title(
            f"{cat_col} 비율",
            color=TEXT, fontsize=12, pad=10, fontweight="bold",
        )
        ax2.legend(
            [str(v)[:12] for v in grp.index.tolist()],
            loc="lower center",
            fontsize=7,
            labelcolor=TEXT,
            facecolor=CARD_BG,
            edgecolor=BORDER,
            bbox_to_anchor=(0.5, -0.12),
            ncol=2,
        )
    ax2.set_facecolor(CARD_BG)

    # ── 하단 출처 ─────────────────────────────────────────
    fig.text(0.5, 0.02,
             f"AI챔피언 데이터분석 인포그래픽  |  데이터 {len(df):,}건",
             ha="center", fontsize=8, color="#6e7681")

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=BG, edgecolor="none")
    buf.seek(0)
    plt.close("all")
    return buf.read()


# ── 5. HTML 정책 제안 리포트 ─────────────────────────────────
def generate_html_policy(
    client: genai.Client,
    model: str,
    df: pd.DataFrame,
    profile: DataProfile,
    problem: Problem,
) -> bytes:
    # 실제 통계 계산
    stat_lines = []
    for num_col in profile.numeric_cols[:4]:
        stat_lines.append(
            f"- {num_col}: 합계={df[num_col].sum():,.1f}, "
            f"평균={df[num_col].mean():.2f}, "
            f"최대={df[num_col].max():,.1f}"
        )
    for cat_col in profile.categorical_cols[:3]:
        top = df[cat_col].value_counts().head(5)
        stat_lines.append(
            f"- {cat_col} 상위: " +
            ", ".join(f"{k}({v:,}건)" for k, v in top.items())
        )

    prompt = f"""다음 데이터 분석 결과를 바탕으로 정책 제안 HTML 리포트를 제작하라.

[시나리오]
{problem.scenario}

[데이터 분석 통계 (실제 계산값 - 반드시 본문에 포함)]
{chr(10).join(stat_lines)}

[필수 포함 요소]
{chr(10).join(f"- {r}" for r in problem.requirements)}

[기술 요건]
- 정부 보고서 스타일 (흰 배경, #1a3a5c 포인트 색상)
- 섹션 구성: 요약(Executive Summary) → 현황 분석 → 주요 발견사항 → 정책 제언 3개 이상 → 결론
- 실제 계산된 통계값을 본문에 자연스럽게 포함 (예: "전체 발생건수 X,XXX건 중...")
- 인쇄 최적화 A4 스타일 (max-width: 800px, margin: auto)
- 자동 목차 생성 (앵커 링크)
- 페이지 상단에 발행일 자동 표시 (JS)
- 배경 요약 카드 (주요 수치 하이라이트)"""

    return _llm_html(client, model, prompt).encode("utf-8")


# ── 통합 산출물 생성 ──────────────────────────────────────────
def generate_all_artifacts(
    client: genai.Client,
    model: str,
    df: pd.DataFrame,
    profile: DataProfile,
    problems: list[Problem],
    progress_callback=None,
) -> dict[str, bytes | tuple]:
    """
    Returns dict:
        filename -> bytes  (HTML, PNG, xlsx)
        "vba_code" -> str  (VBA 매크로 코드)
    """
    artifacts: dict = {}

    for i, problem in enumerate(problems):
        if progress_callback:
            progress_callback(i, problem.type_name)

        if problem.type_id == "html_dashboard":
            artifacts[problem.submission_filename] = generate_html_dashboard(
                client, model, df, profile, problem
            )
        elif problem.type_id == "html_filter":
            artifacts[problem.submission_filename] = generate_html_filter(
                client, model, df, profile, problem
            )
        elif problem.type_id == "excel_vba":
            xlsx_bytes, vba_code = generate_excel_vba(
                client, model, df, profile, problem
            )
            # xlsm 파일명이지만 실제 xlsx로 저장 (VBA는 .bas 파일로 별도 제공)
            artifacts[problem.submission_filename] = xlsx_bytes
            artifacts["vba_code"] = vba_code
        elif problem.type_id == "png_infographic":
            artifacts[problem.submission_filename] = generate_png_infographic(
                df, profile, problem
            )
        elif problem.type_id == "html_policy":
            artifacts[problem.submission_filename] = generate_html_policy(
                client, model, df, profile, problem
            )

    if progress_callback:
        progress_callback(len(problems), "완료")

    return artifacts
