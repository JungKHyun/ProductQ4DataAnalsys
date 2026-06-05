"""채점 기준표 Excel 자동 생성 모듈"""
from __future__ import annotations
import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .problem_gen import Problem

# 스타일 상수
_H_FILL = PatternFill("solid", fgColor="1F4E79")
_SH_FILL = PatternFill("solid", fgColor="2E75B6")
_ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_SCORE_FILL = PatternFill("solid", fgColor="FFF2CC")
_PASS_FILL = PatternFill("solid", fgColor="E2EFDA")
_FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")

_H_FONT = Font(color="FFFFFF", bold=True, size=10)
_BOLD = Font(bold=True, size=10)
_NORM = Font(size=10)
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _sc(cell, fill=None, font=None, align=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def generate_grading_rubric(problems: list[Problem]) -> bytes:
    wb = openpyxl.Workbook()

    # ──────────────────────────────────────────────────────────
    # 시트 1: 총점 요약
    # ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "총점요약"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 15

    # 제목
    ws.merge_cells("A1:G1")
    ws["A1"] = "AI챔피언 데이터분석 실습시험 채점표"
    _sc(ws["A1"],
        font=Font(bold=True, size=18, color="1F4E79"),
        align=_center(), fill=PatternFill("solid", fgColor="DEEAF1"))
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    ws["A2"] = "총 배점 500점  |  합격 기준 350점 이상  |  AI 활용 허용 시험  |  제출물: HTML / Excel / PNG"
    _sc(ws["A2"],
        font=Font(size=10, color="444444"),
        align=_center())
    ws.row_dimensions[2].height = 22

    # 수험자 정보 영역
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22
    for r, (label, col) in enumerate([
        ("수험번호", "B"), ("성    명", "D"), ("소 속", "F")
    ], start=4):
        ws.cell(r, 1, label).font = _BOLD
        ws.cell(r, 2 if col == "B" else (4 if col == "D" else 6), "").border = _THIN
    ws["A4"] = "수험번호"
    ws["A4"].font = _BOLD
    ws["B4"].border = _THIN
    ws["C4"] = "성    명"
    ws["C4"].font = _BOLD
    ws["D4"].border = _THIN
    ws["E4"] = "소속기관"
    ws["E4"].font = _BOLD
    ws["F4"].border = _THIN
    ws.merge_cells("F4:G4")

    # 문제 목록 헤더 (row 6)
    r_hdr = 6
    ws.row_dimensions[r_hdr].height = 25
    headers = ["문제번호", "문제 유형", "문제 제목", "제출 파일명", "배점", "취득점수", "비고"]
    for ci, h in enumerate(headers, 1):
        _sc(ws.cell(r_hdr, ci, h),
            fill=_H_FILL, font=_H_FONT, align=_center(), border=_THIN)

    # 문제별 행
    for i, p in enumerate(problems):
        r = r_hdr + 1 + i
        ws.row_dimensions[r].height = 22
        fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        vals = [f"문제 {p.number}", p.type_name, p.title, p.submission_filename, 100, "", ""]
        aligns = [_center(), _left(False), _left(), _left(), _center(), _center(), _center()]
        for ci, (v, al) in enumerate(zip(vals, aligns), 1):
            _sc(ws.cell(r, ci, v), fill=fill, font=_NORM, align=al, border=_THIN)

    # 합계 행
    r_total = r_hdr + 1 + len(problems)
    ws.row_dimensions[r_total].height = 26
    ws.cell(r_total, 1, "합    계").font = _BOLD
    ws.cell(r_total, 5, 500).font = _BOLD
    score_sum_cell = ws.cell(r_total, 6, f"=SUM(F{r_hdr+1}:F{r_total-1})")
    _sc(score_sum_cell, fill=_SCORE_FILL, font=_BOLD, align=_center(), border=_THIN)
    for ci in [1, 2, 3, 4, 5, 7]:
        _sc(ws.cell(r_total, ci), border=_THIN)

    # 합격 여부
    r_pass = r_total + 2
    ws["A" + str(r_pass)] = "합격 여부"
    ws["A" + str(r_pass)].font = _BOLD
    pass_cell = ws.cell(r_pass, 2,
                         f'=IF(F{r_total}>=350,"✅ 합격","❌ 불합격 (재응시 필요)")')
    _sc(pass_cell, fill=_SCORE_FILL, font=Font(bold=True, size=11),
        align=_center(), border=_THIN)
    ws.merge_cells(f"B{r_pass}:D{r_pass}")
    ws["A" + str(r_pass)].border = _THIN

    ws.cell(r_pass + 1, 1, "※ 합격 기준: 500점 만점 중 350점(70%) 이상").font = Font(size=9, color="666666", italic=True)

    # ──────────────────────────────────────────────────────────
    # 시트 2~6: 문제별 세부 채점표
    # ──────────────────────────────────────────────────────────
    for p in problems:
        ws2 = wb.create_sheet(f"문제{p.number}채점")
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 8
        ws2.column_dimensions["C"].width = 8
        ws2.column_dimensions["D"].width = 10
        ws2.column_dimensions["E"].width = 18
        ws2.column_dimensions["F"].width = 45

        # 문제 헤더
        ws2.merge_cells("A1:F1")
        ws2["A1"] = f"문제 {p.number}  |  [{p.type_name}]  {p.title}"
        _sc(ws2["A1"],
            font=Font(bold=True, size=13, color="1F4E79"),
            align=_center(),
            fill=PatternFill("solid", fgColor="DEEAF1"))
        ws2.row_dimensions[1].height = 32

        # 시나리오
        ws2.merge_cells("A2:F2")
        ws2["A2"] = f"[시나리오] {p.scenario}"
        _sc(ws2["A2"],
            font=Font(size=9, color="444444", italic=True),
            align=Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws2.row_dimensions[2].height = 65

        # 제출 파일 정보
        ws2["A3"] = "제출 파일명"
        ws2["A3"].font = _BOLD
        ws2["B3"] = p.submission_filename
        ws2["B3"].font = Font(bold=True, color="CC0000", size=11)
        ws2.merge_cells("B3:F3")
        ws2.row_dimensions[3].height = 20

        # 채점 헤더 (row 5)
        r_h = 5
        ws2.row_dimensions[r_h].height = 25
        hdrs2 = ["채점 항목", "배점", "체크", "취득점수", "비고", "세부 채점 기준"]
        for ci, h in enumerate(hdrs2, 1):
            _sc(ws2.cell(r_h, ci, h),
                fill=_SH_FILL, font=_H_FONT, align=_center(), border=_THIN)

        # 채점 항목
        for j, ri in enumerate(p.rubric):
            row_idx = r_h + 1 + j
            ws2.row_dimensions[row_idx].height = 32
            fill = _ALT_FILL if j % 2 == 0 else _WHITE_FILL

            _sc(ws2.cell(row_idx, 1, ri.criterion),
                fill=fill, font=_NORM, align=_left(), border=_THIN)
            _sc(ws2.cell(row_idx, 2, ri.score),
                fill=fill, font=_BOLD, align=_center(), border=_THIN)
            _sc(ws2.cell(row_idx, 3, "□"),
                fill=fill, font=Font(size=12), align=_center(), border=_THIN)
            score_formula = f'=IF(C{row_idx}="☑",B{row_idx},0)'
            _sc(ws2.cell(row_idx, 4, score_formula),
                fill=_SCORE_FILL, font=_BOLD, align=_center(), border=_THIN)
            _sc(ws2.cell(row_idx, 5, ""),
                fill=fill, font=_NORM, align=_center(), border=_THIN)
            _sc(ws2.cell(row_idx, 6, ri.description),
                fill=fill, font=_NORM, align=_left(), border=_THIN)

        # 소계
        r_sub = r_h + 1 + len(p.rubric)
        ws2.row_dimensions[r_sub].height = 26
        ws2.merge_cells(f"A{r_sub}:B{r_sub}")
        ws2[f"A{r_sub}"] = "소    계"
        ws2[f"A{r_sub}"].font = _BOLD
        ws2[f"A{r_sub}"].alignment = _center()
        ws2[f"A{r_sub}"].border = _THIN
        _sc(ws2.cell(r_sub, 4, f"=SUM(D{r_h+1}:D{r_sub-1})"),
            fill=_SCORE_FILL, font=Font(bold=True, size=12), align=_center(), border=_THIN)
        ws2.cell(r_sub, 2, sum(ri.score for ri in p.rubric)).border = _THIN
        for ci in [3, 5, 6]:
            ws2.cell(r_sub, ci).border = _THIN

        # 채점 안내
        guide_row = r_sub + 2
        ws2.cell(guide_row, 1, "📌 채점 방법").font = _BOLD
        ws2.merge_cells(f"B{guide_row}:F{guide_row}")
        ws2.cell(guide_row, 2,
                  "제출 파일을 열어 각 항목 확인 후 C열의 □ → ☑ 변경 시 점수 자동 계산됩니다.").font = Font(size=9, italic=True)
        ws2.cell(guide_row, 2).alignment = _left(False)

        # 조건부 채점 예시
        tip_row = guide_row + 1
        type_tips = {
            "html_dashboard": "HTML 파일 브라우저 열기 → KPI 카드 수치 확인 → 차트 렌더링 확인 → 필터 동작 테스트",
            "html_filter": "HTML 파일 열기 → 검색창 입력 테스트 → 드롭다운 필터 동작 → 건수 표시 확인",
            "excel_vba": "xlsx 열기 → 원본데이터/집계요약/자동보고서 시트 확인 → vba_code.bas 내용 검토",
            "png_infographic": "PNG 이미지 열기 → KPI 수치 정확성 → 차트 포함 여부 → 가독성 평가",
            "html_policy": "HTML 열기 → 섹션 구성(요약/분석/제언) 확인 → 통계값 정확성 → 정책 제언 구체성",
        }
        tip = type_tips.get(p.type_id, "파일 열어 채점 항목별 확인")
        ws2.merge_cells(f"B{tip_row}:F{tip_row}")
        ws2.cell(tip_row, 1, "📋 체크 포인트").font = Font(size=9, bold=True, color="2E75B6")
        ws2.cell(tip_row, 2, tip).font = Font(size=9, color="2E75B6")
        ws2.cell(tip_row, 2).alignment = _left(False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
